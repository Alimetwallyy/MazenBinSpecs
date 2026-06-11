import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ================= Page setup =================
st.set_page_config(
    page_title="Bin Divider Specification Generator",
    page_icon="📦",
    layout="wide",
)

st.markdown(
    """
    <style>
        .main .block-container { max-width: 1100px; }
        div[data-testid="stMetricValue"] { font-size: 1.4rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("📦 Bin Divider Specification Generator")
st.caption(
    "Build a central Bin Box Library once, assign bin types to groups, "
    "preview the result live, then export a clean Excel spec."
)

# ================= Session state =================
DEFAULTS = {
    "groups": [],
    "bin_library": {},   # {bin_id: {fields}}
    "next_bin_id": 1,
    "next_group_id": 1,
}
for key, value in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = value

# ================= Constants =================
EXPORT_COLUMNS = [
    "Group Name", "Floor", "Mod", "Depth", "Start Aisle", "End Aisle", "# of Aisles", "# of Bays",
    "Total # of Shelves per Bay", "Bay Design", "Bin Box Type", "Depth (mm)",
    "Height (mm)", "Width (mm)", "Lip (cm)", "# of Shelves per Bay",
    "Qty bins per Shelf", "Qty Per Bay", "Total Quantity", "UT",
    "Bin Gross CBM", "Bin Net CBM",
]

# Columns that describe the group (merged in the Excel export)
GROUP_MERGE_COLS = 9


# ================= Helpers =================
def safe_float(x, default=0.0):
    try:
        return float(x)
    except (TypeError, ValueError):
        return default


def safe_int(x, default=0):
    return int(safe_float(x, default))


def calculate_fields(group_data: dict, bin_data: dict) -> dict:
    """Derive computed bin fields for a (group, bin) pairing.

    Domain rules preserved from the original tool:
      - # of Aisles  = End Aisle - Start Aisle + 1
      - Qty Per Bay  = shelves per bay * bins per shelf
      - Total Qty    = Qty Per Bay * # of Bays
      - Gross CBM    = depth * height * width / 1,000,000
      - Net CBM      = Gross CBM * UT
    """
    g, b = group_data, bin_data
    out = {**b}

    shelves_per_bay = safe_int(b.get("# of Shelves per Bay", 1), 1)
    qty_per_shelf = safe_int(b.get("Qty bins per Shelf", 1), 1)
    ut = safe_float(b.get("UT", 0.0))
    depth_mm = safe_float(b.get("Depth (mm)", 0.0))
    height_mm = safe_float(b.get("Height (mm)", 0.0))
    width_mm = safe_float(b.get("Width (mm)", 0.0))
    lip_cm = safe_float(b.get("Lip (cm)", 0.0))
    start_aisle = safe_int(g.get("Start Aisle", 1), 1)
    end_aisle = safe_int(g.get("End Aisle", 1), 1)
    bays = safe_int(g.get("# of Bays", 1), 1)

    out["Lip (cm)"] = "-" if lip_cm == 0 else round(lip_cm, 2)
    out["# of Aisles"] = end_aisle - start_aisle + 1
    out["Qty Per Bay"] = shelves_per_bay * qty_per_shelf
    out["Total Quantity"] = out["Qty Per Bay"] * bays
    out["Bin Gross CBM"] = round((depth_mm * height_mm * width_mm) / 1_000_000, 4)
    out["Bin Net CBM"] = round(out["Bin Gross CBM"] * ut, 4)
    return out


def sync_bin_keys_with_library():
    """Drop references to bins that no longer exist in the library."""
    valid_ids = set(st.session_state.bin_library.keys())
    for grp in st.session_state.groups:
        grp["bin_keys"] = [k for k in grp.get("bin_keys", []) if k in valid_ids]


def rerun():
    try:
        st.rerun()
    except AttributeError:  # older Streamlit
        st.experimental_rerun()


def bin_label(data: dict, fallback: str) -> str:
    return (data.get("Bin Box Type") or "").strip() or fallback


def build_spec_dataframe(groups: list):
    """Build the full spec DataFrame and the per-group row counts (for merging)."""
    rows = []
    group_row_counts = []
    library = st.session_state.bin_library

    for group in groups:
        group_data = group["group_data"]
        added = 0
        for k in group.get("bin_keys", []):
            base_bin = library.get(k)
            if base_bin is None:
                continue
            calc = calculate_fields(group_data, base_bin)
            row = {**group_data, **calc}
            rows.append({c: row.get(c, None) for c in EXPORT_COLUMNS})
            added += 1
        group_row_counts.append(added)

    df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
    return df, group_row_counts


# ================= Excel export =================
def generate_excel(groups: list) -> bytes:
    df, group_row_counts = build_spec_dataframe(groups)

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Bin Box"

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(r)
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Merge the group-level columns across each group's rows
    current_row = 2
    for row_count in group_row_counts:
        if row_count > 0:
            for col_idx in range(1, GROUP_MERGE_COLS + 1):
                ws.merge_cells(
                    start_row=current_row, start_column=col_idx,
                    end_row=current_row + row_count - 1, end_column=col_idx,
                )
                ws.cell(row=current_row, column=col_idx).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            current_row += row_count

    # Auto-size columns for readability
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        max_len = len(str(col_name))
        for value in df[col_name].astype(str):
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ================= Sidebar: summary =================
def render_summary():
    df, _ = build_spec_dataframe(st.session_state.groups)
    total_qty = int(df["Total Quantity"].fillna(0).sum()) if not df.empty else 0
    total_net = float(df["Bin Net CBM"].fillna(0).sum()) if not df.empty else 0.0

    with st.sidebar:
        st.header("Summary")
        st.metric("Bin types", len(st.session_state.bin_library))
        st.metric("Groups", len(st.session_state.groups))
        st.metric("Total bins", f"{total_qty:,}")
        st.metric("Total Net CBM", f"{total_net:,.3f}")
        st.divider()
        if st.button("🧹 Clear all data", use_container_width=True):
            st.session_state.groups = []
            st.session_state.bin_library = {}
            st.session_state.next_bin_id = 1
            st.session_state.next_group_id = 1
            rerun()


# ================= Tab 1: Bin Library =================
def add_bin(seed: dict | None = None):
    new_id = f"bin_{st.session_state.next_bin_id}"
    st.session_state.next_bin_id += 1
    st.session_state.bin_library[new_id] = seed.copy() if seed else {
        "Bin Box Type": "",
        "Depth (mm)": 0.0,
        "Height (mm)": 0.0,
        "Width (mm)": 0.0,
        "Lip (cm)": 0.0,
        "# of Shelves per Bay": 1,
        "Qty bins per Shelf": 1,
        "UT": 0.0,
    }
    return new_id


def render_bin_library():
    st.subheader("Bin Box Library")
    st.write("Define each bin type once. They become available to every group.")

    if st.button("➕ Add new bin type"):
        add_bin()
        rerun()

    if not st.session_state.bin_library:
        st.info("No bin types yet. Add one to get started.")
        return

    for bin_id, data in list(st.session_state.bin_library.items()):
        label = bin_label(data, bin_id)
        with st.expander(label, expanded=False):
            c1, c2 = st.columns(2)
            with c1:
                data["Bin Box Type"] = st.text_input(
                    "Bin Box Type", value=data.get("Bin Box Type", ""), key=f"name_{bin_id}"
                )
                data["Depth (mm)"] = st.number_input(
                    "Depth (mm)", min_value=0.0, value=safe_float(data.get("Depth (mm)")),
                    step=0.1, key=f"depth_{bin_id}",
                )
                data["Height (mm)"] = st.number_input(
                    "Height (mm)", min_value=0.0, value=safe_float(data.get("Height (mm)")),
                    step=0.1, key=f"height_{bin_id}",
                )
                has_lip = st.checkbox(
                    "Has lip?", value=safe_float(data.get("Lip (cm)")) > 0, key=f"lip_chk_{bin_id}"
                )
            with c2:
                data["Width (mm)"] = st.number_input(
                    "Width (mm)", min_value=0.0, value=safe_float(data.get("Width (mm)")),
                    step=0.1, key=f"width_{bin_id}",
                )
                data["Lip (cm)"] = (safe_float(data["Height (mm)"]) * 0.2 / 10) if has_lip else 0.0
                st.number_input(
                    "Lip (cm) — auto", value=safe_float(data["Lip (cm)"]),
                    disabled=True, key=f"lip_val_{bin_id}",
                )
                data["# of Shelves per Bay"] = int(st.number_input(
                    "# of Shelves per Bay", min_value=1,
                    value=safe_int(data.get("# of Shelves per Bay", 1), 1),
                    step=1, key=f"shelves_{bin_id}",
                ))
                data["Qty bins per Shelf"] = int(st.number_input(
                    "Qty bins per Shelf", min_value=1,
                    value=safe_int(data.get("Qty bins per Shelf", 1), 1),
                    step=1, key=f"qty_{bin_id}",
                ))
                data["UT"] = st.number_input(
                    "UT (0-1)", min_value=0.0, max_value=1.0,
                    value=min(max(safe_float(data.get("UT")), 0.0), 1.0),
                    step=0.01, key=f"ut_{bin_id}",
                )

            # Live per-bin volume readout
            gross = (safe_float(data["Depth (mm)"]) * safe_float(data["Height (mm)"])
                     * safe_float(data["Width (mm)"])) / 1_000_000
            mc1, mc2 = st.columns(2)
            mc1.metric("Gross CBM", f"{gross:.4f}")
            mc2.metric("Net CBM", f"{gross * safe_float(data.get('UT')):.4f}")

            b1, b2 = st.columns(2)
            if b1.button("📄 Duplicate", key=f"dup_{bin_id}"):
                seed = data.copy()
                seed["Bin Box Type"] = f"{label} (copy)"
                add_bin(seed)
                rerun()
            if b2.button("🗑️ Delete", key=f"del_{bin_id}"):
                del st.session_state.bin_library[bin_id]
                sync_bin_keys_with_library()
                rerun()


# ================= Tab 2: Groups =================
def add_group(seed: dict | None = None):
    st.session_state.next_group_id += 1
    if seed:
        st.session_state.groups.append({
            "group_data": seed["group_data"].copy(),
            "bin_keys": list(seed.get("bin_keys", [])),
            "finalized": False,
        })
    else:
        st.session_state.groups.append({
            "group_data": {
                "Group Name": "", "Floor": "", "Mod": "", "Depth": "",
                "Start Aisle": 1, "End Aisle": 1, "# of Bays": 1,
                "Total # of Shelves per Bay": 1, "Bay Design": "",
            },
            "bin_keys": [],
            "finalized": False,
        })


def render_groups():
    st.subheader("Manage Groups")
    if not st.session_state.bin_library:
        st.warning("Add at least one bin type in the Bin Library tab first.")

    if st.button("➕ Add new group"):
        add_group()
        rerun()

    if not st.session_state.groups:
        st.info("No groups yet. Add one to start assigning bins.")
        return

    available_ids = list(st.session_state.bin_library.keys())
    labels = {k: bin_label(st.session_state.bin_library[k], k) for k in available_ids}

    delete_idx = None
    for group_idx, group in enumerate(st.session_state.groups):
        gd = group["group_data"]
        title = gd["Group Name"] or "Untitled"
        state = "✅ Finalized" if group["finalized"] else "✏️ Editing"
        with st.expander(f"Group {group_idx + 1}: {title} ({state})", expanded=not group["finalized"]):
            if not group["finalized"]:
                c1, c2 = st.columns(2)
                with c1:
                    gd["Group Name"] = st.text_input("Group Name", value=gd["Group Name"], key=f"gname_{group_idx}")
                    gd["Floor"] = st.text_input("Floor", value=gd["Floor"], key=f"gflr_{group_idx}")
                    gd["Mod"] = st.text_input("Mod", value=gd["Mod"], key=f"gmod_{group_idx}")
                    gd["Depth"] = st.text_input("Depth", value=gd["Depth"], key=f"gdepth_{group_idx}")
                with c2:
                    gd["Start Aisle"] = int(st.number_input(
                        "Start Aisle", min_value=1, value=safe_int(gd["Start Aisle"], 1),
                        step=1, key=f"gstart_{group_idx}"))
                    gd["End Aisle"] = int(st.number_input(
                        "End Aisle", min_value=1, value=safe_int(gd["End Aisle"], 1),
                        step=1, key=f"gend_{group_idx}"))
                    gd["# of Bays"] = int(st.number_input(
                        "# of Bays", min_value=1, value=safe_int(gd["# of Bays"], 1),
                        step=1, key=f"gbays_{group_idx}"))
                    gd["Total # of Shelves per Bay"] = int(st.number_input(
                        "Total # of Shelves per Bay", min_value=1,
                        value=safe_int(gd["Total # of Shelves per Bay"], 1),
                        step=1, key=f"gshelves_{group_idx}"))
                    gd["Bay Design"] = st.text_input("Bay Design", value=gd["Bay Design"], key=f"gbay_{group_idx}")

                if gd["End Aisle"] < gd["Start Aisle"]:
                    st.error("End Aisle must be greater than or equal to Start Aisle.")

                default_vals = [k for k in group.get("bin_keys", []) if k in available_ids]
                group["bin_keys"] = st.multiselect(
                    "Bin Box Types for this group", options=available_ids, default=default_vals,
                    format_func=lambda k: labels[k], key=f"binsel_{group_idx}",
                )
                if not group["bin_keys"]:
                    st.caption("⚠️ This group has no bins assigned and won't appear in the export.")

                b1, b2, b3 = st.columns(3)
                if b1.button("✅ Finalize", key=f"gfin_{group_idx}"):
                    group["finalized"] = True
                    rerun()
                if b2.button("📄 Duplicate", key=f"gdup_{group_idx}"):
                    add_group(group)
                    rerun()
                if b3.button("🗑️ Delete", key=f"gdel_{group_idx}"):
                    delete_idx = group_idx
            else:
                summary = ", ".join(labels.get(k, k) for k in group["bin_keys"]) or "no bins"
                st.write(f"**Bins:** {summary}")
                e1, e2 = st.columns(2)
                if e1.button("✏️ Edit", key=f"gedit_{group_idx}"):
                    group["finalized"] = False
                    rerun()
                if e2.button("🗑️ Delete", key=f"gdel_fin_{group_idx}"):
                    delete_idx = group_idx

    if delete_idx is not None:
        st.session_state.groups.pop(delete_idx)
        rerun()


# ================= Tab 3: Preview & Export =================
def render_preview_export():
    st.subheader("Preview & Export")
    sync_bin_keys_with_library()
    df, _ = build_spec_dataframe(st.session_state.groups)

    if df.empty:
        st.info("Nothing to preview yet. Add bin types and groups, then assign bins to groups.")
        return

    m1, m2, m3 = st.columns(3)
    m1.metric("Rows", len(df))
    m2.metric("Total bins", f"{int(df['Total Quantity'].fillna(0).sum()):,}")
    m3.metric("Total Net CBM", f"{float(df['Bin Net CBM'].fillna(0).sum()):,.3f}")

    st.dataframe(df, use_container_width=True, hide_index=True)

    excel_data = generate_excel(st.session_state.groups)
    st.download_button(
        label="⬇️ Download Excel file",
        data=excel_data,
        file_name="Bin_Divider_Specs.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ================= Layout =================
sync_bin_keys_with_library()
render_summary()

tab_library, tab_groups, tab_export = st.tabs(
    ["📦 Bin Library", "🏢 Groups", "📊 Preview & Export"]
)
with tab_library:
    render_bin_library()
with tab_groups:
    render_groups()
with tab_export:
    render_preview_export()
